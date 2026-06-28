"""
Extract station-level L-D 00-24 average ratings from Buró data.
"""

import csv
import os

import openpyxl
import xlrd

DATA = os.path.dirname(os.path.abspath(__file__))


def parse_ld_sheet_2425(ws):
    """Parse RatxHoraxLD-type sheet. Stations in col 1, 24-hr avg in last col (30).
    Returns dict band -> [(station, avg)]."""
    results = {"AM": [], "FM": []}
    band = None
    header_lines = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None or not isinstance(v, str):
            continue
        v = v.strip()
        if v.startswith("CUADRO "):
            if "| AM |" in v:
                band = "AM"
            elif "| FM |" in v:
                band = "FM"
            else:
                band = None
            header_lines = 0
            continue
        if v in ("RADIOS",) or v.startswith("RADIOS "):
            header_lines = 1
            continue
        if header_lines > 0:
            header_lines -= 1
            continue
        if band is None:
            continue
        if v in (
            "SUB TOTAL AM",
            "SUB TOTAL FM",
            "RATING GENERAL",
            "AM SIN IDENTIFICAR",
            "FM SIN IDENTIFICAR",
            "ZAPPING AM",
            "ZAPPING FM",
            "ZAPPING EN AM",
            "ZAPPING EN FM",
            "OTRAS AM",
            "OTRAS FM",
            "FM SIN IDENTIFICAR",
        ):
            continue
        avg = ws.cell(r, 30).value
        if avg is None:
            continue
        try:
            avg = float(avg)
        except ValueError, TypeError:
            continue
        results[band].append((v, avg))
    return results


def extract(filepath, sheet="RatxHoraxLD"):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    return parse_ld_sheet_2425(wb[sheet])


def extract_2022():
    res = {"AM": [], "FM": []}
    wb = xlrd.open_workbook(f"{DATA}/2022/Rating x Hora L-D.xls", formatting_info=False)
    sh = wb.sheet_by_index(0)
    band = "AM"
    for r in range(7, sh.nrows):  # data starts at row 7
        name = str(sh.cell(r, 0).value or "").strip()
        if not name:
            continue
        if name == "SUB TOTAL AM":
            band = "FM"
            continue
        if name in (
            "SUB TOTAL FM",
            "TOTAL GENERAL",
            "RATING GENERAL",
            "AM SIN IDENTIFICAR",
            "FM SIN IDENTIFICAR",
            "ZAPPING EN AM",
            "ZAPPING EN FM",
            "OTRAS AM",
            "OTRAS FM",
            "FM SIN IDENTIFICAR",
            "RADIOS",
        ):
            continue
        avg = sh.cell(r, 29).value
        if avg is None:
            continue
        try:
            avg = float(avg)
        except ValueError, TypeError:
            continue
        res[band].append((name, avg))
    return res


def print_res(year, data):
    print(f"\n{'=' * 50}\n{year}\n{'=' * 50}")
    for band in ["AM", "FM"]:
        sts = data.get(band, [])
        if sts:
            print(f"\n  {band}:")
            for s, a in sorted(sts, key=lambda x: -x[1]):
                print(f"    {s:30s} {a:.4f}")
            print(f"    {'---':30s} ----")
            print(f"    {'SUB TOTAL':30s} {sum(a for _, a in sts):.4f}")


def save_csv(all_data):
    with open(f"{DATA}/station_ratings.csv", "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["station", "band", "year", "avg_rating"])
        for year, data in all_data.items():
            for band, sts in data.items():
                for s, a in sts:
                    w.writerow([s, band, year, f"{a:.4f}"])
    print(f"\nSaved {DATA}/station_ratings.csv")


if __name__ == "__main__":
    all_data = {}
    for yr in ["2017", "2018", "2019"]:
        d = extract(f"{DATA}/{yr}.xlsx")
        print_res(yr, d)
        all_data[yr] = d
    d2021 = extract(f"{DATA}/2021.xlsx", "RatxHoraxLD")
    print_res("2021", d2021)
    all_data["2021"] = d2021
    d2022 = extract_2022()
    print_res("2022", d2022)
    all_data["2022"] = d2022
    save_csv(all_data)
